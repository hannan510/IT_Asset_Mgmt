"""
IT Asset Management Application
A Flask app for tracking IT assets with a dashboard.
Run: python app.py  ->  http://localhost:7077
"""
import os
import io
import csv
import sqlite3
from datetime import datetime, date
from flask import (Flask, render_template, request, redirect, url_for, flash,
                   jsonify, Response, send_file)

app = Flask(__name__)
app.secret_key = "change-this-secret-key"
# Store the DB next to app.py (i.e. in your project folder). Override with
# the ASSET_DB env var. Note: on Windows, if the project folder path is very
# deep (>260 chars, e.g. under AppData), SQLite may fail with "unable to open
# database file" — in that case set ASSET_DB to a shorter path.
DB_PATH = os.environ.get(
    "ASSET_DB", os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets.db")
)

# Currency displayed throughout the app.
CURRENCY = "AED"

# Columns used for exports (db_field, header label).
EXPORT_COLUMNS = [
    ("asset_tag", "Asset Tag"),
    ("name", "Name"),
    ("serial_number", "Serial Number"),
    ("category", "Category"),
    ("status", "Status"),
    ("assigned_to", "Assigned To"),
    ("employee_id", "Employee ID"),
    ("phone", "Phone"),
    ("email", "Email"),
    ("location", "Location"),
    ("purchase_cost", "Purchase Cost (AED)"),
    ("purchase_date", "Purchase Date"),
    ("warranty_end", "Warranty End"),
    ("notes", "Notes"),
]

# ---------------------------------------------------------------------------
# Active Directory (on-prem LDAP) integration
# ---------------------------------------------------------------------------
# Configure via environment variables (see ad_config.example.txt). For a quick
# demo without a real domain, set AD_MOCK=1 to use sample users.
AD_CONFIG = {
    "server": os.environ.get("AD_SERVER", ""),          # e.g. ldap://dc01.corp.local
    "port": int(os.environ.get("AD_PORT", "0")) or None,  # default 389, or 636 for SSL
    "use_ssl": os.environ.get("AD_USE_SSL", "").lower() in ("1", "true", "yes"),
    "base_dn": os.environ.get("AD_BASE_DN", ""),         # e.g. DC=corp,DC=local
    "bind_dn": os.environ.get("AD_BIND_DN", ""),         # service account (UPN or DN)
    "bind_password": os.environ.get("AD_BIND_PASSWORD", ""),
    "mock": os.environ.get("AD_MOCK", "").lower() in ("1", "true", "yes"),
}
AD_CONFIG["enabled"] = bool(AD_CONFIG["server"]) or AD_CONFIG["mock"]

_AD_MOCK_USERS = [
    {"name": "Alice Johnson", "employee_id": "EMP-1001", "phone": "+971 50 123 4567", "email": "alice.johnson@corp.local"},
    {"name": "Bob Lee", "employee_id": "EMP-1002", "phone": "+971 52 234 5678", "email": "bob.lee@corp.local"},
    {"name": "Carol White", "employee_id": "EMP-1003", "phone": "+971 55 345 6789", "email": "carol.white@corp.local"},
    {"name": "Dan Brown", "employee_id": "EMP-1004", "phone": "+971 56 456 7890", "email": "dan.brown@corp.local"},
    {"name": "Eve Davis", "employee_id": "EMP-1005", "phone": "+971 54 567 8901", "email": "eve.davis@corp.local"},
    {"name": "Frank Miller", "employee_id": "EMP-1006", "phone": "+971 50 678 9012", "email": "frank.miller@corp.local"},
]


def ad_search(term, limit=20):
    """Search Active Directory for users matching `term`.
    Returns a list of {name, employee_id, phone, email}.
    Uses mock data if AD_MOCK is set; otherwise queries LDAP via ldap3.
    """
    term = (term or "").strip()
    if not term:
        return []
    if AD_CONFIG["mock"]:
        t = term.lower()
        return [u for u in _AD_MOCK_USERS
                if t in u["name"].lower() or t in u["employee_id"].lower()
                or t in u["email"].lower()][:limit]
    if not AD_CONFIG["enabled"]:
        raise RuntimeError(
            "Active Directory is not configured. Set AD_SERVER, AD_BASE_DN, "
            "AD_BIND_DN and AD_BIND_PASSWORD (see ad_config.example.txt), or set "
            "AD_MOCK=1 to try it with sample data.")
    from ldap3 import Server, Connection, ALL, SUBTREE
    from ldap3.utils.conv import escape_filter_chars
    safe = escape_filter_chars(term)
    server = Server(AD_CONFIG["server"], port=AD_CONFIG["port"],
                    use_ssl=AD_CONFIG["use_ssl"], get_info=ALL)
    conn = Connection(server, user=AD_CONFIG["bind_dn"],
                      password=AD_CONFIG["bind_password"], auto_bind=True)
    flt = ("(&(objectCategory=person)(objectClass=user)"
           "(|(displayName=*{0}*)(sAMAccountName=*{0}*)(mail=*{0}*)"
           "(employeeID=*{0}*)))".format(safe))
    conn.search(AD_CONFIG["base_dn"], flt, search_scope=SUBTREE,
                attributes=["displayName", "employeeID", "telephoneNumber",
                            "mobile", "mail", "sAMAccountName"],
                size_limit=limit)
    out = []
    for e in conn.entries:
        def g(attr):
            v = getattr(e, attr, None)
            return str(v) if v else ""
        out.append({
            "name": g("displayName") or g("sAMAccountName"),
            "employee_id": g("employeeID"),
            "phone": g("telephoneNumber") or g("mobile"),
            "email": g("mail"),
        })
    conn.unbind()
    return out


@app.route("/api/ad/search")
def api_ad_search():
    configured = AD_CONFIG["enabled"]
    try:
        users = ad_search(request.args.get("q", ""))
        return jsonify(ok=True, configured=configured, users=users)
    except Exception as exc:  # surfaced to the UI as a friendly message
        return jsonify(ok=False, configured=configured, error=str(exc), users=[])

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
# Default categories. Users can also type a brand-new category on the asset
# form (e.g. "Keyboard") and it is remembered automatically afterwards.
CATEGORIES = ["Laptop", "Desktop", "Monitor", "Server", "Network", "Phone",
              "Printer", "Tablet", "Keyboard", "Mouse", "Headset",
              "Docking Station", "Webcam", "Scanner", "Router", "Switch",
              "UPS", "Peripheral", "Software License", "Other"]
STATUSES = ["In Use", "In Storage", "In Repair", "Retired", "Lost"]


def _all_categories():
    """Default categories plus any custom ones already saved in the database."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT DISTINCT category FROM assets WHERE category IS NOT NULL "
            "AND category != ''"
        ).fetchall()
    except sqlite3.OperationalError:
        rows = []
    conn.close()
    cats = list(CATEGORIES)
    for r in rows:
        if r["category"] not in cats:
            cats.append(r["category"])
    return cats


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS assets (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_tag     TEXT    NOT NULL UNIQUE,
            name          TEXT    NOT NULL,
            category      TEXT    NOT NULL,
            status        TEXT    NOT NULL,
            manufacturer  TEXT,
            model         TEXT,
            serial_number TEXT,
            assigned_to   TEXT,
            employee_id   TEXT,
            phone         TEXT,
            email         TEXT,
            location      TEXT,
            purchase_date TEXT,
            purchase_cost REAL    DEFAULT 0,
            warranty_end  TEXT,
            notes         TEXT,
            created_at    TEXT
        )
        """
    )
    conn.commit()

    # --- Lightweight migration ---------------------------------------------
    # If the database was created by an older version, add any columns that are
    # missing (e.g. employee_id, phone) so older databases keep working.
    existing = {row[1] for row in conn.execute("PRAGMA table_info(assets)").fetchall()}
    # Every non-id column the app expects. Columns that already exist are
    # skipped, so this is safe to run on any older database. (NOT NULL is
    # omitted here because ALTER TABLE ADD COLUMN can't add a NOT NULL column
    # to a table that already has rows.)
    wanted = {
        "asset_tag": "TEXT", "name": "TEXT", "category": "TEXT", "status": "TEXT",
        "manufacturer": "TEXT", "model": "TEXT", "serial_number": "TEXT",
        "assigned_to": "TEXT", "employee_id": "TEXT", "phone": "TEXT", "email": "TEXT",
        "location": "TEXT", "purchase_date": "TEXT", "purchase_cost": "REAL DEFAULT 0",
        "warranty_end": "TEXT", "notes": "TEXT", "created_at": "TEXT",
    }
    for col, decl in wanted.items():
        if col not in existing:
            conn.execute("ALTER TABLE assets ADD COLUMN %s %s" % (col, decl))
    conn.commit()

    # Seed demo data if empty
    count = conn.execute("SELECT COUNT(*) AS c FROM assets").fetchone()["c"]
    if count == 0:
        seed = [
            ("LT-0001", "Dev Laptop 01", "Laptop", "In Use", "Dell", "Latitude 7440",
             "SN-DL-7440-A1", "Alice Johnson", "EMP-1001", "+971 50 123 4567", "HQ - Floor 3", "2024-03-15", 5325.00, "2027-03-15", "Engineering"),
            ("LT-0002", "Design MacBook", "Laptop", "In Use", "Apple", "MacBook Pro 16",
             "SN-APL-MBP-22", "Bob Lee", "EMP-1002", "+971 52 234 5678", "HQ - Floor 2", "2023-11-02", 10645.00, "2026-11-02", "Design team"),
            ("DT-0010", "Finance Desktop", "Desktop", "In Use", "HP", "EliteDesk 800",
             "SN-HP-800-77", "Carol White", "EMP-1003", "+971 55 345 6789", "HQ - Floor 1", "2022-06-20", 3600.00, "2025-06-20", ""),
            ("SV-0003", "App Server Prod", "Server", "In Use", "Dell", "PowerEdge R750",
             "SN-PE-R750-X", "", "", "", "Datacenter Rack 4", "2023-01-10", 28650.00, "2028-01-10", "Production"),
            ("SV-0004", "Backup Server", "Server", "In Repair", "Dell", "PowerEdge R650",
             "SN-PE-R650-Y", "", "", "", "Datacenter Rack 5", "2022-09-05", 22770.00, "2027-09-05", "RAID issue"),
            ("MN-0021", "Monitor 27 4K", "Monitor", "In Use", "LG", "27UN880",
             "SN-LG-27-01", "Alice Johnson", "EMP-1001", "+971 50 123 4567", "HQ - Floor 3", "2024-03-15", 1395.00, "2027-03-15", ""),
            ("MN-0022", "Monitor 24", "Monitor", "In Storage", "Dell", "P2422H",
             "SN-DL-24-09", "", "", "", "Storage Room B", "2023-05-12", 770.00, "2026-05-12", "Spare"),
            ("NW-0007", "Core Switch", "Network", "In Use", "Cisco", "Catalyst 9300",
             "SN-CSC-9300-Z", "", "", "", "Datacenter Rack 1", "2021-12-01", 16530.00, "2026-12-01", ""),
            ("PH-0031", "iPhone 15", "Phone", "In Use", "Apple", "iPhone 15 Pro",
             "SN-APL-IP15-3", "Dan Brown", "EMP-1004", "+971 56 456 7890", "Remote", "2024-01-20", 4405.00, "2026-01-20", "Sales"),
            ("PR-0012", "Office Printer", "Printer", "In Use", "Brother", "MFC-L8900",
             "SN-BR-L8900-2", "", "", "", "HQ - Floor 2", "2022-02-15", 1985.00, "2025-02-15", "Shared"),
            ("LT-0005", "Old Laptop", "Laptop", "Retired", "Lenovo", "ThinkPad T480",
             "SN-LN-T480-5", "", "", "", "Storage Room B", "2019-08-10", 4040.00, "2022-08-10", "EOL"),
            ("SW-0100", "Adobe CC License", "Software License", "In Use", "Adobe", "Creative Cloud",
             "KEY-ADB-2024", "Bob Lee", "EMP-1002", "+971 52 234 5678", "Cloud", "2024-01-01", 2205.00, "2025-01-01", "Annual"),
        ]
        conn.executemany(
            """INSERT INTO assets
               (asset_tag,name,category,status,manufacturer,model,serial_number,
                assigned_to,employee_id,phone,location,purchase_date,purchase_cost,
                warranty_end,notes,created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [row + (datetime.now().isoformat(),) for row in seed],
        )
        conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def dashboard():
    # Optional dashboard filters.
    f_category = request.args.get("category", "")
    f_location = request.args.get("location", "")

    where = " WHERE 1=1"
    params = []
    if f_category:
        where += " AND category = ?"
        params.append(f_category)
    if f_location:
        where += " AND location = ?"
        params.append(f_location)

    conn = get_db()
    total = conn.execute("SELECT COUNT(*) AS c FROM assets" + where, params).fetchone()["c"]
    total_value = conn.execute(
        "SELECT COALESCE(SUM(purchase_cost),0) AS v FROM assets" + where, params
    ).fetchone()["v"]
    in_use = conn.execute(
        "SELECT COUNT(*) AS c FROM assets" + where + " AND status='In Use'", params
    ).fetchone()["c"]
    in_repair = conn.execute(
        "SELECT COUNT(*) AS c FROM assets" + where + " AND status='In Repair'", params
    ).fetchone()["c"]

    by_category = conn.execute(
        "SELECT category, COUNT(*) AS c FROM assets" + where +
        " GROUP BY category ORDER BY c DESC", params
    ).fetchall()
    by_status = conn.execute(
        "SELECT status, COUNT(*) AS c FROM assets" + where +
        " GROUP BY status ORDER BY c DESC", params
    ).fetchall()

    # Warranty expiring within 90 days (or already expired)
    today = date.today().isoformat()
    expiring = conn.execute(
        "SELECT * FROM assets" + where +
        " AND warranty_end IS NOT NULL AND warranty_end != ''"
        " AND warranty_end <= date('now','+90 day')"
        " ORDER BY warranty_end ASC LIMIT 10", params
    ).fetchall()
    recent = conn.execute(
        "SELECT * FROM assets" + where + " ORDER BY id DESC LIMIT 5", params
    ).fetchall()

    locations = conn.execute(
        "SELECT DISTINCT location FROM assets WHERE location IS NOT NULL "
        "AND location != '' ORDER BY location"
    ).fetchall()
    conn.close()

    stats = {
        "total": total,
        "total_value": total_value,
        "in_use": in_use,
        "in_repair": in_repair,
    }
    return render_template(
        "dashboard.html",
        stats=stats,
        by_category=[dict(r) for r in by_category],
        by_status=[dict(r) for r in by_status],
        expiring=expiring,
        recent=recent,
        today=today,
        categories=_all_categories(),
        locations=[r["location"] for r in locations],
        sel_category=f_category,
        sel_location=f_location,
    )


def _filtered_assets(q="", category="", status="", location=""):
    """Return asset rows matching the given filters."""
    conn = get_db()
    sql = "SELECT * FROM assets WHERE 1=1"
    params = []
    if q:
        sql += (" AND (asset_tag LIKE ? OR name LIKE ? OR assigned_to LIKE ?"
                " OR employee_id LIKE ? OR phone LIKE ? OR serial_number LIKE ?)")
        params += [f"%{q}%"] * 6
    if category:
        sql += " AND category = ?"
        params.append(category)
    if status:
        sql += " AND status = ?"
        params.append(status)
    if location:
        sql += " AND location = ?"
        params.append(location)
    sql += " ORDER BY id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def _all_locations():
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT location FROM assets WHERE location IS NOT NULL "
        "AND location != '' ORDER BY location"
    ).fetchall()
    conn.close()
    return [r["location"] for r in rows]


@app.route("/assets")
def asset_list():
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "")
    status = request.args.get("status", "")
    assets = _filtered_assets(q, category, status)
    return render_template(
        "assets.html", assets=assets, categories=_all_categories(), statuses=STATUSES,
        q=q, sel_category=category, sel_status=status,
    )


# ---------------------------------------------------------------------------
# Exports (CSV / Excel / PDF) — honour the current search & filters
# ---------------------------------------------------------------------------
def _export_rows():
    return _filtered_assets(
        request.args.get("q", "").strip(),
        request.args.get("category", ""),
        request.args.get("status", ""),
        request.args.get("location", ""),
    )


@app.route("/export/csv")
def export_csv():
    rows = _export_rows()
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for r in rows:
        writer.writerow([r[field] for field, _ in EXPORT_COLUMNS])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets.csv"},
    )


@app.route("/export/xlsx")
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Excel export needs openpyxl. Run: pip install openpyxl", "error")
        return redirect(url_for("asset_list"))
    rows = _export_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    headers = [label for _, label in EXPORT_COLUMNS]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="left")
    for r in rows:
        ws.append([r[field] for field, _ in EXPORT_COLUMNS])
    # Auto-ish column widths
    for col_idx, (field, label) in enumerate(EXPORT_COLUMNS, start=1):
        width = max(len(label), 12)
        for r in rows:
            val = r[field]
            if val is not None:
                width = max(width, min(len(str(val)) + 2, 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio, as_attachment=True, download_name="assets.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/pdf")
def export_pdf():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
    except ImportError:
        flash("PDF export needs reportlab. Run: pip install reportlab", "error")
        return redirect(url_for("asset_list"))
    rows = _export_rows()
    # A trimmed column set so it fits a landscape page.
    pdf_cols = [
        ("asset_tag", "Tag"), ("name", "Name"), ("serial_number", "Serial"),
        ("category", "Category"), ("status", "Status"), ("assigned_to", "Assigned"),
        ("employee_id", "Emp ID"), ("phone", "Phone"), ("purchase_cost", "Cost (AED)"),
    ]
    data = [[label for _, label in pdf_cols]]
    for r in rows:
        line = []
        for field, _ in pdf_cols:
            val = r[field]
            if field == "purchase_cost":
                val = "{:,.0f}".format(float(val or 0))
            line.append("" if val is None else str(val))
        data.append(line)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4),
                            leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("IT Asset Inventory", styles["Title"]),
        Paragraph("Generated " + datetime.now().strftime("%Y-%m-%d %H:%M") +
                  "  ·  " + str(len(rows)) + " asset(s)", styles["Normal"]),
        Spacer(1, 12),
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="assets.pdf",
                     mimetype="application/pdf")


@app.route("/assets/new", methods=["GET", "POST"])
def asset_new():
    if request.method == "POST":
        data = _form_data()
        conn = get_db()
        try:
            conn.execute(
                """INSERT INTO assets
                   (asset_tag,name,category,status,manufacturer,model,serial_number,
                    assigned_to,employee_id,phone,email,location,purchase_date,purchase_cost,
                    warranty_end,notes,created_at)
                   VALUES (:asset_tag,:name,:category,:status,:manufacturer,:model,:serial_number,
                    :assigned_to,:employee_id,:phone,:email,:location,:purchase_date,:purchase_cost,
                    :warranty_end,:notes,:created_at)""",
                {**data, "created_at": datetime.now().isoformat()},
            )
            conn.commit()
            flash("Asset created.", "success")
            return redirect(url_for("asset_list"))
        except sqlite3.IntegrityError:
            flash("Asset tag already exists. Use a unique tag.", "error")
        finally:
            conn.close()
    return render_template(
        "asset_form.html", asset=None, categories=_all_categories(), statuses=STATUSES,
        action="new", ad_enabled=AD_CONFIG["enabled"]
    )


@app.route("/assets/<int:asset_id>/edit", methods=["GET", "POST"])
def asset_edit(asset_id):
    conn = get_db()
    asset = conn.execute("SELECT * FROM assets WHERE id=?", (asset_id,)).fetchone()
    if asset is None:
        conn.close()
        flash("Asset not found.", "error")
        return redirect(url_for("asset_list"))
    if request.method == "POST":
        data = _form_data()
        try:
            conn.execute(
                """UPDATE assets SET
                   asset_tag=:asset_tag,name=:name,category=:category,status=:status,
                   manufacturer=:manufacturer,model=:model,serial_number=:serial_number,
                   assigned_to=:assigned_to,employee_id=:employee_id,phone=:phone,email=:email,
                   location=:location,purchase_date=:purchase_date,
                   purchase_cost=:purchase_cost,warranty_end=:warranty_end,notes=:notes
                   WHERE id=:id""",
                {**data, "id": asset_id},
            )
            conn.commit()
            flash("Asset updated.", "success")
            return redirect(url_for("asset_list"))
        except sqlite3.IntegrityError:
            flash("Asset tag already exists. Use a unique tag.", "error")
        finally:
            conn.close()
        return redirect(url_for("asset_edit", asset_id=asset_id))
    conn.close()
    return render_template(
        "asset_form.html", asset=asset, categories=_all_categories(), statuses=STATUSES,
        action="edit", ad_enabled=AD_CONFIG["enabled"]
    )


@app.route("/assets/<int:asset_id>/delete", methods=["POST"])
def asset_delete(asset_id):
    conn = get_db()
    conn.execute("DELETE FROM assets WHERE id=?", (asset_id,))
    conn.commit()
    conn.close()
    flash("Asset deleted.", "success")
    return redirect(url_for("asset_list"))


@app.route("/api/stats")
def api_stats():
    conn = get_db()
    by_category = conn.execute(
        "SELECT category, COUNT(*) AS c FROM assets GROUP BY category"
    ).fetchall()
    by_status = conn.execute(
        "SELECT status, COUNT(*) AS c FROM assets GROUP BY status"
    ).fetchall()
    conn.close()
    return jsonify(
        by_category={r["category"]: r["c"] for r in by_category},
        by_status={r["status"]: r["c"] for r in by_status},
    )


def _form_data():
    def g(k):
        return request.form.get(k, "").strip()
    cost = g("purchase_cost")
    try:
        cost = float(cost) if cost else 0.0
    except ValueError:
        cost = 0.0
    return {
        "asset_tag": g("asset_tag"),
        "name": g("name"),
        "category": g("category") or "Other",
        "status": g("status") or "In Use",
        "manufacturer": g("manufacturer"),
        "model": g("model"),
        "serial_number": g("serial_number"),
        "assigned_to": g("assigned_to"),
        "employee_id": g("employee_id"),
        "phone": g("phone"),
        "email": g("email"),
        "location": g("location"),
        "purchase_date": g("purchase_date"),
        "purchase_cost": cost,
        "warranty_end": g("warranty_end"),
        "notes": g("notes"),
    }


@app.template_filter("money")
def money(v):
    try:
        return "{} {:,.0f}".format(CURRENCY, float(v or 0))
    except (ValueError, TypeError):
        return "{} 0".format(CURRENCY)


if __name__ == "__main__":
    init_db()
    print("IT Asset Manager running at http://localhost:7077")
    app.run(host="0.0.0.0", port=7077, debug=True)
